from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
import string,io
def write():
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'a1'
    img = Image('.tmp/1.png')
    ws.add_image(img,'B1')
    wb.save('.tmp/1.xlsx')


import PIL
class SheetImages:
    """Loads all images in a sheet"""

    def __init__(self, ws):
        self._images = {}
        for image in ws._images:
            row = image.anchor._from.row + 1
            col = string.ascii_uppercase[image.anchor._from.col]
            self._images[f'{col}{row}'] = image._data

    def exists(self, cell):
        return cell in self._images

    def get(self, cell):
        if not self.exists(cell):
            return None
        else:
            image = io.BytesIO(self._images[cell]())
            return PIL.Image.open(image)    

 
def read():
    wb = load_workbook('.tmp/1.xlsx')
    ws = wb.active
    print(ws['A1'].value)
    pass

    si = SheetImages(ws)
    img = si.get('B1')
    if img != None:
        img.save('.tmp/~1.png')


#write()
read()